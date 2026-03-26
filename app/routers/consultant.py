"""Consultant dashboard API — client management, batch processing, deadlines, revenue, templates."""

import csv
import io
import uuid
from dataclasses import asdict
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_

from app.core.database import get_db
from app.core.security import get_current_consultant
from app.core.config import get_settings
from app.models.user import User, UserRole
from app.models.client import Client, ClientStatus, EntityType
from app.models.tax_filing import TaxFiling, FilingStatus
from app.models.document import Document, DocumentStatus
from app.models.invoice import Invoice, InvoiceStatus
from app.models.portal_token import PortalToken
from app.models.document_verification import DocumentVerification
from app.models.spt_template import SPTTemplate
from app.models.consent import ClientConsent, ConsentType, ConsentStatus
from app.models.audit_log import AuditLog
from app.models.whatsapp_message import WhatsAppMessage
from app.services.tax_calendar import ObligationType, get_client_deadlines
from app.services.batch_processor import process_batch_ocr, batch_generate_spt
from app.services.nudge_engine import send_document_request

router = APIRouter(prefix="/consultant", tags=["consultant"])
settings = get_settings()


# ──────────────────────────────────────────────
# Schemas
# ──────────────────────────────────────────────

class ClientCreate(BaseModel):
    name: str
    npwp: str | None = None
    nik: str | None = None
    email: str | None = None
    phone: str | None = None
    entity_type: str = "orang_pribadi"
    is_pkp: bool = False
    ptkp_status: str | None = None
    tax_obligations: list[str] | None = None
    fee_monthly: float | None = None
    notes: str | None = None


class ClientUpdate(BaseModel):
    name: str | None = None
    npwp: str | None = None
    nik: str | None = None
    email: str | None = None
    phone: str | None = None
    entity_type: str | None = None
    is_pkp: bool | None = None
    ptkp_status: str | None = None
    tax_obligations: list[str] | None = None
    fee_monthly: float | None = None
    notes: str | None = None
    status: str | None = None


class ClientOut(BaseModel):
    id: str
    name: str
    npwp: str | None
    phone: str | None
    entity_type: str
    status: str
    is_pkp: bool
    ptkp_status: str | None
    tax_obligations: list | None
    fee_monthly: float | None

    model_config = {"from_attributes": True}


class InvoiceCreate(BaseModel):
    client_id: str
    amount: float
    description: str | None = None
    due_date: str | None = None
    items: list[dict] | None = None


class InvoiceOut(BaseModel):
    id: str
    client_id: str
    client_name: str | None = None
    amount: float
    description: str | None
    status: str
    due_date: str | None
    created_at: str
    paid_at: str | None


class TemplateCreate(BaseModel):
    name: str
    filing_type: str
    description: str | None = None
    template_data: dict


class TemplateOut(BaseModel):
    id: str
    name: str
    filing_type: str
    description: str | None
    created_at: str


class BatchOCRRequest(BaseModel):
    documents: list[dict]  # [{id, file_url, doc_type, mime_type}]


class BatchSPTRequest(BaseModel):
    clients: list[dict]  # [{client_id, bukti_potong_extractions, user_data}]


class DocRequestPayload(BaseModel):
    client_id: str
    doc_types: list[str]
    masa: int
    tahun: int


class FilingStatusUpdate(BaseModel):
    status: str


class DocumentVerifyPayload(BaseModel):
    verified: bool
    notes: str | None = None


# ──────────────────────────────────────────────
# Client Management
# ──────────────────────────────────────────────

@router.get("/clients")
async def list_clients(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: str = Query(""),
    status: str = Query(""),
    entity_type: str = Query(""),
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    """List clients with pagination, search, and filtering."""
    query = select(Client).where(Client.consultant_id == user.id)
    if search:
        search_term = f"%{search}%"
        query = query.where(
            or_(
                Client.name.ilike(search_term),
                Client.npwp.ilike(search_term),
                Client.email.ilike(search_term),
                Client.phone.ilike(search_term),
            )
        )
    if status:
        query = query.where(Client.status == ClientStatus(status))
    if entity_type:
        query = query.where(Client.entity_type == EntityType(entity_type))
    count_q = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_q)).scalar() or 0
    offset = (page - 1) * page_size
    result = await db.execute(query.order_by(Client.name).offset(offset).limit(page_size))
    clients = result.scalars().all()
    return {
        "items": [_client_out(c) for c in clients],
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size,
    }


@router.post("/clients", response_model=ClientOut)
async def create_client(
    data: ClientCreate,
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    client = Client(
        consultant_id=user.id,
        name=data.name,
        npwp=data.npwp,
        nik=data.nik,
        email=data.email,
        phone=data.phone,
        entity_type=EntityType(data.entity_type),
        is_pkp=data.is_pkp,
        ptkp_status=data.ptkp_status,
        tax_obligations=data.tax_obligations,
        fee_monthly=data.fee_monthly,
        notes=data.notes,
    )
    db.add(client)
    await db.commit()
    await db.refresh(client)
    await _log_activity(db, user.id, "client_created", f"Created client {client.name}", client_id=client.id)
    return _client_out(client)


@router.get("/clients/{client_id}", response_model=ClientOut)
async def get_client(
    client_id: str,
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    client = await _get_consultant_client(client_id, user.id, db)
    return _client_out(client)


@router.put("/clients/{client_id}", response_model=ClientOut)
async def update_client(
    client_id: str,
    data: ClientUpdate,
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    """Update client details."""
    client = await _get_consultant_client(client_id, user.id, db)
    update_data = data.model_dump(exclude_unset=True)
    if "entity_type" in update_data and update_data["entity_type"] is not None:
        update_data["entity_type"] = EntityType(update_data["entity_type"])
    if "status" in update_data and update_data["status"] is not None:
        update_data["status"] = ClientStatus(update_data["status"])
    for key, value in update_data.items():
        setattr(client, key, value)
    await db.commit()
    await db.refresh(client)
    await _log_activity(db, user.id, "client_updated", f"Updated client {client.name}", client_id=client.id)
    return _client_out(client)


@router.delete("/clients/{client_id}")
async def delete_client(
    client_id: str,
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    client = await _get_consultant_client(client_id, user.id, db)
    client.status = ClientStatus.INACTIVE
    await db.commit()
    await _log_activity(db, user.id, "client_deactivated", f"Deactivated client {client.name}", client_id=client.id)
    return {"status": "deactivated"}


# ──────────────────────────────────────────────
# Bulk CSV/Excel Import
# ──────────────────────────────────────────────

@router.post("/clients/import")
async def import_clients(
    file: UploadFile = File(...),
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    """Bulk import clients from CSV or Excel file.

    CSV columns: name, npwp, nik, email, phone, entity_type, is_pkp, ptkp_status, tax_obligations, fee_monthly, notes
    """
    if not file.filename:
        raise HTTPException(400, "No file provided")

    content = await file.read()
    imported = []
    errors = []

    if file.filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    elif file.filename.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [str(cell.value).strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row) if i < len(headers)}
                rows.append(row_dict)
        except ImportError:
            raise HTTPException(400, "openpyxl is required for Excel import")
    else:
        raise HTTPException(400, "Unsupported file format. Use .csv or .xlsx")

    for idx, row in enumerate(rows, start=2):
        name = row.get("name", "").strip()
        if not name:
            errors.append({"row": idx, "error": "Missing name"})
            continue

        try:
            entity_type_val = row.get("entity_type", "orang_pribadi").strip() or "orang_pribadi"
            is_pkp_raw = row.get("is_pkp", "false").strip().lower()
            is_pkp = is_pkp_raw in ("true", "1", "yes", "ya")
            obligations_raw = row.get("tax_obligations", "").strip()
            tax_obligations = [o.strip() for o in obligations_raw.split(",") if o.strip()] if obligations_raw else None
            fee_raw = row.get("fee_monthly", "").strip()
            fee_monthly = float(fee_raw) if fee_raw else None

            client = Client(
                consultant_id=user.id,
                name=name,
                npwp=row.get("npwp", "").strip() or None,
                nik=row.get("nik", "").strip() or None,
                email=row.get("email", "").strip() or None,
                phone=row.get("phone", "").strip() or None,
                entity_type=EntityType(entity_type_val),
                is_pkp=is_pkp,
                ptkp_status=row.get("ptkp_status", "").strip() or None,
                tax_obligations=tax_obligations,
                fee_monthly=fee_monthly,
                notes=row.get("notes", "").strip() or None,
            )
            db.add(client)
            imported.append(name)
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})

    if imported:
        await db.commit()
        await _log_activity(db, user.id, "clients_imported", f"Imported {len(imported)} clients via {file.filename}")

    return {
        "imported": len(imported),
        "errors": errors,
        "total_rows": len(rows),
    }


# ──────────────────────────────────────────────
# Deadlines
# ──────────────────────────────────────────────

@router.get("/deadlines")
async def get_all_deadlines(
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    """Get upcoming deadlines across all clients."""
    result = await db.execute(
        select(Client).where(
            Client.consultant_id == user.id,
            Client.status == ClientStatus.ACTIVE,
        )
    )
    clients = result.scalars().all()

    all_deadlines = []
    for client in clients:
        obligations = [
            ObligationType(o) for o in (client.tax_obligations or [])
            if o in [e.value for e in ObligationType]
        ]
        if not obligations:
            continue

        deadlines = get_client_deadlines(obligations=obligations, months_ahead=2)
        for dl in deadlines:
            all_deadlines.append({
                "client_id": client.id,
                "client_name": client.name,
                "obligation": dl.obligation.value,
                "label": dl.label,
                "deadline_date": str(dl.deadline_date),
                "payment_deadline": str(dl.payment_deadline),
                "status": dl.status.value,
                "days_remaining": dl.days_remaining,
                "penalty_amount": dl.penalty_amount,
            })

    all_deadlines.sort(key=lambda d: d["deadline_date"])
    return {
        "total_clients": len(clients),
        "total_deadlines": len(all_deadlines),
        "overdue": sum(1 for d in all_deadlines if d["status"] == "overdue"),
        "due_soon": sum(1 for d in all_deadlines if d["status"] == "due_soon"),
        "deadlines": all_deadlines,
    }


# ──────────────────────────────────────────────
# Filing Kanban Board
# ──────────────────────────────────────────────

@router.get("/filings/board")
async def get_filing_board(
    tax_year: int | None = Query(None),
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    """Get filings grouped by status for Kanban board view."""
    # Get all client IDs for this consultant
    client_q = select(Client.id).where(Client.consultant_id == user.id)
    client_result = await db.execute(client_q)
    client_ids = [row[0] for row in client_result.all()]

    if not client_ids:
        return {"columns": {s.value: [] for s in FilingStatus}, "total": 0}

    query = (
        select(TaxFiling, Client.name.label("client_name"))
        .join(Client, TaxFiling.client_id == Client.id)
        .where(TaxFiling.client_id.in_(client_ids))
    )
    if tax_year:
        query = query.where(TaxFiling.tax_year == tax_year)

    result = await db.execute(query.order_by(TaxFiling.created_at.desc()))
    rows = result.all()

    columns: dict[str, list] = {s.value: [] for s in FilingStatus}
    for filing, client_name in rows:
        item = {
            "id": filing.id,
            "client_id": filing.client_id,
            "client_name": client_name,
            "filing_type": filing.filing_type.value,
            "tax_year": filing.tax_year,
            "tax_month": filing.tax_month,
            "status": filing.status.value,
            "deadline": str(filing.deadline) if filing.deadline else None,
            "created_at": str(filing.created_at),
        }
        columns[filing.status.value].append(item)

    return {
        "columns": columns,
        "total": len(rows),
    }


@router.patch("/filings/{filing_id}/status")
async def update_filing_status(
    filing_id: str,
    data: FilingStatusUpdate,
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    """Move a filing to a new status (Kanban drag-and-drop)."""
    # Verify the filing belongs to one of this consultant's clients
    result = await db.execute(
        select(TaxFiling)
        .join(Client, TaxFiling.client_id == Client.id)
        .where(TaxFiling.id == filing_id, Client.consultant_id == user.id)
    )
    filing = result.scalar_one_or_none()
    if not filing:
        raise HTTPException(404, "Filing not found")

    old_status = filing.status.value
    filing.status = FilingStatus(data.status)
    if data.status == "filed" and not filing.filed_at:
        filing.filed_at = datetime.now(timezone.utc)
    await db.commit()
    await _log_activity(
        db, user.id, "filing_status_changed",
        f"Filing {filing_id} status: {old_status} -> {data.status}",
        client_id=filing.client_id,
    )
    return {"id": filing_id, "status": data.status}


# ──────────────────────────────────────────────
# Document Verification Workflow
# ──────────────────────────────────────────────

@router.get("/documents/pending-review")
async def get_pending_documents(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    """List documents awaiting consultant review."""
    client_q = select(Client.id).where(Client.consultant_id == user.id)
    client_result = await db.execute(client_q)
    client_ids = [row[0] for row in client_result.all()]

    if not client_ids:
        return {"items": [], "total": 0, "page": page, "page_size": page_size, "total_pages": 0}

    query = (
        select(Document, Client.name.label("client_name"))
        .join(TaxFiling, Document.filing_id == TaxFiling.id)
        .join(Client, TaxFiling.client_id == Client.id)
        .where(
            TaxFiling.client_id.in_(client_ids),
            Document.status == DocumentStatus.EXTRACTED,
        )
    )
    count_q = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_q)).scalar() or 0
    offset = (page - 1) * page_size
    result = await db.execute(query.order_by(Document.created_at.desc()).offset(offset).limit(page_size))
    rows = result.all()

    items = []
    for doc, client_name in rows:
        items.append({
            "id": doc.id,
            "filing_id": doc.filing_id,
            "client_name": client_name,
            "doc_type": doc.doc_type.value,
            "file_name": doc.file_name,
            "ocr_confidence": doc.ocr_confidence,
            "status": doc.status.value,
            "created_at": str(doc.created_at),
        })

    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size,
    }


@router.post("/documents/{document_id}/verify")
async def verify_document(
    document_id: str,
    data: DocumentVerifyPayload,
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    """Verify or reject a document after review."""
    # Find the document and ensure it belongs to this consultant's client
    result = await db.execute(
        select(Document)
        .join(TaxFiling, Document.filing_id == TaxFiling.id)
        .join(Client, TaxFiling.client_id == Client.id)
        .where(Document.id == document_id, Client.consultant_id == user.id)
    )
    doc = result.scalar_one_or_none()
    if not doc:
        raise HTTPException(404, "Document not found")

    if data.verified:
        doc.status = DocumentStatus.VERIFIED
    else:
        doc.status = DocumentStatus.REJECTED

    # Record the verification
    verification = DocumentVerification(
        document_id=document_id,
        consultant_id=user.id,
        verified=data.verified,
        notes=data.notes,
        verified_at=datetime.now(timezone.utc),
    )
    db.add(verification)
    await db.commit()
    await _log_activity(
        db, user.id, "document_verified" if data.verified else "document_rejected",
        f"Document {document_id} {'verified' if data.verified else 'rejected'}",
    )
    return {"id": document_id, "status": doc.status.value, "verified": data.verified}


# ──────────────────────────────────────────────
# Portal Token Management
# ──────────────────────────────────────────────

@router.post("/clients/{client_id}/portal-link")
async def create_portal_link(
    client_id: str,
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    """Generate a portal access token/link for a client."""
    client = await _get_consultant_client(client_id, user.id, db)
    token_value = uuid.uuid4().hex
    portal_token = PortalToken(
        client_id=client.id,
        token=token_value,
        created_by=user.id,
        created_at=datetime.now(timezone.utc),
    )
    db.add(portal_token)
    await db.commit()
    await db.refresh(portal_token)

    portal_url = f"{settings.frontend_url}/portal?token={token_value}"
    await _log_activity(db, user.id, "portal_link_created", f"Portal link created for {client.name}", client_id=client.id)
    return {
        "client_id": client.id,
        "client_name": client.name,
        "token": token_value,
        "portal_url": portal_url,
        "created_at": str(portal_token.created_at),
    }


# ──────────────────────────────────────────────
# Invoice CRUD
# ──────────────────────────────────────────────

@router.post("/invoices", response_model=InvoiceOut)
async def create_invoice(
    data: InvoiceCreate,
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    """Create an invoice for a client."""
    client = await _get_consultant_client(data.client_id, user.id, db)
    invoice = Invoice(
        consultant_id=user.id,
        client_id=client.id,
        amount=data.amount,
        description=data.description,
        due_date=datetime.fromisoformat(data.due_date) if data.due_date else None,
        items=data.items,
    )
    db.add(invoice)
    await db.commit()
    await db.refresh(invoice)
    await _log_activity(db, user.id, "invoice_created", f"Invoice created for {client.name}: Rp {data.amount:,.0f}", client_id=client.id)
    return _invoice_out(invoice, client.name)


@router.get("/invoices")
async def list_invoices(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    status: str = Query(""),
    client_id: str = Query(""),
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    """List invoices with pagination and filters."""
    query = (
        select(Invoice, Client.name.label("client_name"))
        .join(Client, Invoice.client_id == Client.id)
        .where(Invoice.consultant_id == user.id)
    )
    if status:
        query = query.where(Invoice.status == InvoiceStatus(status))
    if client_id:
        query = query.where(Invoice.client_id == client_id)

    count_q = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_q)).scalar() or 0
    offset = (page - 1) * page_size
    result = await db.execute(query.order_by(Invoice.created_at.desc()).offset(offset).limit(page_size))
    rows = result.all()

    items = [_invoice_out(inv, client_name) for inv, client_name in rows]
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size,
    }


@router.patch("/invoices/{invoice_id}/status")
async def update_invoice_status(
    invoice_id: str,
    data: FilingStatusUpdate,
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    """Update invoice status (draft -> sent -> paid / cancelled)."""
    result = await db.execute(
        select(Invoice).where(Invoice.id == invoice_id, Invoice.consultant_id == user.id)
    )
    invoice = result.scalar_one_or_none()
    if not invoice:
        raise HTTPException(404, "Invoice not found")

    old_status = invoice.status.value
    invoice.status = InvoiceStatus(data.status)
    if data.status == "paid" and not invoice.paid_at:
        invoice.paid_at = datetime.now(timezone.utc)
    await db.commit()
    await _log_activity(
        db, user.id, "invoice_status_changed",
        f"Invoice {invoice_id} status: {old_status} -> {data.status}",
        client_id=invoice.client_id,
    )
    return {"id": invoice_id, "status": data.status}


# ──────────────────────────────────────────────
# SPT Templates
# ──────────────────────────────────────────────

@router.get("/templates")
async def list_templates(
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    """List all SPT templates created by this consultant."""
    result = await db.execute(
        select(SPTTemplate).where(SPTTemplate.consultant_id == user.id).order_by(SPTTemplate.created_at.desc())
    )
    templates = result.scalars().all()
    return [
        TemplateOut(
            id=t.id,
            name=t.name,
            filing_type=t.filing_type,
            description=t.description,
            created_at=str(t.created_at),
        )
        for t in templates
    ]


@router.post("/templates", response_model=TemplateOut)
async def create_template(
    data: TemplateCreate,
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    """Create a reusable SPT template."""
    template = SPTTemplate(
        consultant_id=user.id,
        name=data.name,
        filing_type=data.filing_type,
        description=data.description,
        template_data=data.template_data,
    )
    db.add(template)
    await db.commit()
    await db.refresh(template)
    await _log_activity(db, user.id, "template_created", f"Template created: {data.name}")
    return TemplateOut(
        id=template.id,
        name=template.name,
        filing_type=template.filing_type,
        description=template.description,
        created_at=str(template.created_at),
    )


@router.delete("/templates/{template_id}")
async def delete_template(
    template_id: str,
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    """Delete an SPT template."""
    result = await db.execute(
        select(SPTTemplate).where(SPTTemplate.id == template_id, SPTTemplate.consultant_id == user.id)
    )
    template = result.scalar_one_or_none()
    if not template:
        raise HTTPException(404, "Template not found")
    await db.delete(template)
    await db.commit()
    await _log_activity(db, user.id, "template_deleted", f"Template deleted: {template.name}")
    return {"status": "deleted"}


# ──────────────────────────────────────────────
# Multi-Year History
# ──────────────────────────────────────────────

@router.get("/clients/{client_id}/history")
async def get_client_history(
    client_id: str,
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    """Get multi-year filing history for a client."""
    client = await _get_consultant_client(client_id, user.id, db)

    result = await db.execute(
        select(TaxFiling)
        .where(TaxFiling.client_id == client.id)
        .order_by(TaxFiling.tax_year.desc(), TaxFiling.tax_month.desc().nullslast())
    )
    filings = result.scalars().all()

    # Group by year
    by_year: dict[int, list] = {}
    for f in filings:
        year_data = {
            "id": f.id,
            "filing_type": f.filing_type.value,
            "status": f.status.value,
            "tax_month": f.tax_month,
            "gross_income": float(f.gross_income) if f.gross_income else None,
            "tax_due": float(f.tax_due) if f.tax_due else None,
            "filed_at": str(f.filed_at) if f.filed_at else None,
        }
        by_year.setdefault(f.tax_year, []).append(year_data)

    # Calculate year-over-year summaries
    year_summaries = []
    sorted_years = sorted(by_year.keys(), reverse=True)
    for year in sorted_years:
        year_filings = by_year[year]
        total_income = sum(f["gross_income"] or 0 for f in year_filings)
        total_tax = sum(f["tax_due"] or 0 for f in year_filings)
        year_summaries.append({
            "year": year,
            "filings": year_filings,
            "total_filings": len(year_filings),
            "total_gross_income": total_income,
            "total_tax_due": total_tax,
            "filed_count": sum(1 for f in year_filings if f["status"] == "filed"),
        })

    return {
        "client_id": client.id,
        "client_name": client.name,
        "years": year_summaries,
        "total_filings": len(filings),
    }


# ──────────────────────────────────────────────
# Client Timeline
# ──────────────────────────────────────────────

@router.get("/clients/{client_id}/timeline")
async def get_client_timeline(
    client_id: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    """Get combined timeline of all activity for a client."""
    client = await _get_consultant_client(client_id, user.id, db)
    events = []

    # Filings
    filings_result = await db.execute(
        select(TaxFiling).where(TaxFiling.client_id == client.id).order_by(TaxFiling.created_at.desc())
    )
    for f in filings_result.scalars().all():
        events.append({
            "type": "filing",
            "id": f.id,
            "title": f"Filing {f.filing_type.value} - {f.tax_year}" + (f"/{f.tax_month}" if f.tax_month else ""),
            "status": f.status.value,
            "timestamp": str(f.created_at),
        })
        if f.filed_at:
            events.append({
                "type": "filing_submitted",
                "id": f.id,
                "title": f"Filed {f.filing_type.value} - {f.tax_year}" + (f"/{f.tax_month}" if f.tax_month else ""),
                "timestamp": str(f.filed_at),
            })

    # Documents
    doc_result = await db.execute(
        select(Document)
        .join(TaxFiling, Document.filing_id == TaxFiling.id)
        .where(TaxFiling.client_id == client.id)
        .order_by(Document.created_at.desc())
    )
    for d in doc_result.scalars().all():
        events.append({
            "type": "document",
            "id": d.id,
            "title": f"Document uploaded: {d.file_name}",
            "status": d.status.value,
            "timestamp": str(d.created_at),
        })

    # WhatsApp messages
    wa_result = await db.execute(
        select(WhatsAppMessage)
        .where(WhatsAppMessage.client_id == client.id)
        .order_by(WhatsAppMessage.created_at.desc())
        .limit(50)
    )
    for msg in wa_result.scalars().all():
        events.append({
            "type": "whatsapp",
            "id": msg.id,
            "title": f"WhatsApp: {msg.direction} - {msg.message_type}",
            "timestamp": str(msg.created_at),
        })

    # Audit log entries for this client
    audit_result = await db.execute(
        select(AuditLog)
        .where(AuditLog.client_id == client.id)
        .order_by(AuditLog.created_at.desc())
        .limit(50)
    )
    for log in audit_result.scalars().all():
        events.append({
            "type": "activity",
            "id": log.id,
            "title": log.description,
            "action": log.action,
            "timestamp": str(log.created_at),
        })

    # Sort all events by timestamp descending
    events.sort(key=lambda e: e["timestamp"], reverse=True)

    # Paginate
    total = len(events)
    offset = (page - 1) * page_size
    paginated = events[offset:offset + page_size]

    return {
        "client_id": client.id,
        "client_name": client.name,
        "items": paginated,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size,
    }


# ──────────────────────────────────────────────
# Batch Processing
# ──────────────────────────────────────────────

@router.post("/batch/ocr")
async def batch_ocr(
    data: BatchOCRRequest,
    user: User = Depends(get_current_consultant),
):
    """Process multiple documents through OCR in parallel."""
    result = await process_batch_ocr(data.documents)
    return asdict(result)


@router.post("/batch/spt")
async def batch_spt(
    data: BatchSPTRequest,
    user: User = Depends(get_current_consultant),
):
    """Generate SPT for multiple clients in parallel."""
    results = await batch_generate_spt(data.clients)
    return {"results": results}


# ──────────────────────────────────────────────
# Document Requests
# ──────────────────────────────────────────────

@router.post("/request-documents")
async def request_documents(
    data: DocRequestPayload,
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    """Send document collection request to client via WhatsApp."""
    client = await _get_consultant_client(data.client_id, user.id, db)
    phone = client.whatsapp_id or client.phone
    if not phone:
        raise HTTPException(400, "Client has no WhatsApp number")

    result = await send_document_request(
        phone=phone,
        client_name=client.name,
        doc_types=data.doc_types,
        masa=data.masa,
        tahun=data.tahun,
    )
    await _log_activity(db, user.id, "document_request_sent", f"Document request sent to {client.name}", client_id=client.id)
    return result


# ──────────────────────────────────────────────
# Enhanced Analytics
# ──────────────────────────────────────────────

@router.get("/analytics")
async def get_analytics(
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    """Dashboard analytics for consultant with revenue and document stats."""
    # Client stats
    result = await db.execute(
        select(Client).where(Client.consultant_id == user.id)
    )
    clients = result.scalars().all()
    active = [c for c in clients if c.status == ClientStatus.ACTIVE]
    total_revenue = sum(c.fee_monthly or 0 for c in active)

    entity_breakdown = {}
    for c in active:
        et = c.entity_type.value
        entity_breakdown[et] = entity_breakdown.get(et, 0) + 1

    # Filing stats
    client_ids = [c.id for c in clients]
    filing_stats = {"total": 0, "by_status": {}}
    if client_ids:
        filing_result = await db.execute(
            select(TaxFiling.status, func.count(TaxFiling.id))
            .where(TaxFiling.client_id.in_(client_ids))
            .group_by(TaxFiling.status)
        )
        for status, count in filing_result.all():
            filing_stats["by_status"][status.value] = count
            filing_stats["total"] += count

    # Document stats
    doc_stats = {"total": 0, "by_status": {}}
    if client_ids:
        doc_result = await db.execute(
            select(Document.status, func.count(Document.id))
            .join(TaxFiling, Document.filing_id == TaxFiling.id)
            .where(TaxFiling.client_id.in_(client_ids))
            .group_by(Document.status)
        )
        for status, count in doc_result.all():
            doc_stats["by_status"][status.value] = count
            doc_stats["total"] += count

    # Invoice / revenue stats
    revenue_stats = {"total_invoiced": 0, "total_paid": 0, "total_outstanding": 0}
    invoice_result = await db.execute(
        select(Invoice.status, func.sum(Invoice.amount))
        .where(Invoice.consultant_id == user.id)
        .group_by(Invoice.status)
    )
    for status, total in invoice_result.all():
        amount = float(total) if total else 0
        revenue_stats["total_invoiced"] += amount
        if status == InvoiceStatus.PAID:
            revenue_stats["total_paid"] += amount
        elif status in (InvoiceStatus.SENT, InvoiceStatus.DRAFT):
            revenue_stats["total_outstanding"] += amount

    return {
        "total_clients": len(clients),
        "active_clients": len(active),
        "monthly_revenue": total_revenue,
        "annual_revenue_projection": total_revenue * 12,
        "entity_breakdown": entity_breakdown,
        "pkp_clients": sum(1 for c in active if c.is_pkp),
        "non_pkp_clients": sum(1 for c in active if not c.is_pkp),
        "filing_stats": filing_stats,
        "document_stats": doc_stats,
        "revenue_stats": revenue_stats,
    }


# ──────────────────────────────────────────────
# Activity Log
# ──────────────────────────────────────────────

@router.get("/activity-log")
async def get_activity_log(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    client_id: str = Query(""),
    action: str = Query(""),
    user: User = Depends(get_current_consultant),
    db: AsyncSession = Depends(get_db),
):
    """Get activity log for this consultant."""
    query = select(AuditLog).where(AuditLog.user_id == user.id)
    if client_id:
        query = query.where(AuditLog.client_id == client_id)
    if action:
        query = query.where(AuditLog.action == action)

    count_q = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_q)).scalar() or 0
    offset = (page - 1) * page_size
    result = await db.execute(query.order_by(AuditLog.created_at.desc()).offset(offset).limit(page_size))
    logs = result.scalars().all()

    items = [
        {
            "id": log.id,
            "action": log.action,
            "description": log.description,
            "client_id": log.client_id,
            "created_at": str(log.created_at),
        }
        for log in logs
    ]

    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size,
    }


# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────

async def _get_consultant_client(client_id: str, consultant_id: str, db: AsyncSession) -> Client:
    result = await db.execute(
        select(Client).where(Client.id == client_id, Client.consultant_id == consultant_id)
    )
    client = result.scalar_one_or_none()
    if not client:
        raise HTTPException(404, "Client not found")
    return client


def _client_out(c: Client) -> ClientOut:
    return ClientOut(
        id=c.id,
        name=c.name,
        npwp=c.npwp,
        phone=c.phone,
        entity_type=c.entity_type.value,
        status=c.status.value,
        is_pkp=c.is_pkp,
        ptkp_status=c.ptkp_status,
        tax_obligations=c.tax_obligations,
        fee_monthly=c.fee_monthly,
    )


def _invoice_out(inv: Invoice, client_name: str | None = None) -> InvoiceOut:
    return InvoiceOut(
        id=inv.id,
        client_id=inv.client_id,
        client_name=client_name,
        amount=float(inv.amount),
        description=inv.description,
        status=inv.status.value,
        due_date=str(inv.due_date) if inv.due_date else None,
        created_at=str(inv.created_at),
        paid_at=str(inv.paid_at) if inv.paid_at else None,
    )


async def _log_activity(
    db: AsyncSession,
    user_id: str,
    action: str,
    description: str,
    client_id: str | None = None,
):
    """Record an activity in the audit log."""
    log = AuditLog(
        user_id=user_id,
        action=action,
        description=description,
        client_id=client_id,
        created_at=datetime.now(timezone.utc),
    )
    db.add(log)
    await db.commit()
